"""
Care Provider Finder - England Only (Auto-download latest CQC ODS if older than 31 days)
Handles ODS files using pandas + odf.
"""

import os
import time
import gdown
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# === Settings ===
DATA_DIR = r"C:\Users\murre\Documents\GitHub\care-provider-finder-http\data"
os.makedirs(DATA_DIR, exist_ok=True)

# Latest file ID from Google Drive link (ODS file)
GDRIVE_FILE_ID = "1pwkUrk9iZsCiWiNSiKZjYbqVxdYJoxCi"
CQC_ODS = os.path.join(DATA_DIR, "cqc_care_directory.ods")
OUTPUT_FILE = "Care_Provider_England.xlsx"

# === Step 1: Download latest file if needed ===
download_needed = True
if os.path.exists(CQC_ODS):
    age_days = (time.time() - os.path.getmtime(CQC_ODS)) / (24 * 3600)
    if age_days < 31:
        download_needed = False
        print(f"Using existing CQC ODS ({age_days:.1f} days old)")
    else:
        print(f"CQC ODS older than 31 days ({age_days:.1f} days), downloading new version")
else:
    print("CQC ODS not found, downloading latest version...")

if download_needed:
    url = f"https://drive.google.com/uc?id={GDRIVE_FILE_ID}"
    print(f"Downloading latest CQC ODS to {CQC_ODS}...")
    gdown.download(url, CQC_ODS, quiet=False)
    print("Download complete.")

# === Step 2: Read ODS file ===
try:
    df_cqc = pd.read_excel(CQC_ODS, engine="odf")
    print(f"✅ Loaded {len(df_cqc)} records from CQC ODS")
except Exception as e:
    print(f"❌ Error loading CQC ODS: {e}")
    df_cqc = pd.DataFrame()

# === Step 3: Prepare workbook ===
wb = openpyxl.Workbook()
wb.remove(wb.active)

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def write_dataframe(ws, df, source_desc):
    ws["A1"] = ws.title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=max(4, len(df.columns))
    )
    ws["A2"] = f"Data loaded: {len(df)} records"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A3"] = f"Source: {source_desc}"
    ws["A3"].font = Font(italic=True, size=9, color="999999")

    start_row = 5
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                ws.row_dimensions[start_row].height = 28

    for idx, column in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A6"

# === Step 4: Create sheets ===
sheet_names = [
    "📍 Radius Search", "⭐ Outstanding", "⚠️ At-Risk",
    "🔍 Recent Inspections", "🏢 Large Homes",
    "🆕 New Registrations", "🎯 Dementia Care", "📊 Regional Stats"
]

summary_data = []

for sn in sheet_names:
    ws = wb.create_sheet(sn)
    if sn == "📍 Radius Search" and not df_cqc.empty:
        write_dataframe(ws, df_cqc, CQC_ODS)
        summary_data.append({"Sheet": sn, "Rows": len(df_cqc)})
    else:
        write_dataframe(ws, pd.DataFrame(), "No data")
        summary_data.append({"Sheet": sn, "Rows": 0})

# Combined sheet
combined_ws = wb.create_sheet("📌 Combined")
write_dataframe(combined_ws, df_cqc, "Combined: Radius Search")
summary_data.append({"Sheet": "📌 Combined", "Rows": len(df_cqc)})

# Summary sheet
summary_ws = wb.create_sheet("📄 Summary")
summary_df = pd.DataFrame(summary_data)
write_dataframe(summary_ws, summary_df, "Summary of sheets")

# Color-code summary
for row in range(6, 6 + len(summary_df)):
    cell = summary_ws.cell(row=row, column=2)
    cell.fill = green_fill if cell.value > 0 else red_fill

# Save workbook
wb.save(OUTPUT_FILE)
print(f"\n✅ Excel file created: {OUTPUT_FILE}")
